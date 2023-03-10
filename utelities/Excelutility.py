import openpyxl
from openpyxl.styles import PatternFill


def get_rowcount(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return(sheet.max_row)


def get_columncount(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return(sheet.max_column)


def readData(file,sheetname,rownum,colnum):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return sheet.cell(rownum,colnum).value


def writeData(file,sheetname,rownum,colnum,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    sheet.cell(rownum,colnum).value=data
    workbook.save(file)


def fillGreencolor(file,sheetname,rownum,colnum):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    greenfill=PatternFill(start_color='60b212',
                          end_color='60b212',
                          fill_type='solid')
    sheet.cell(rownum,colnum).fill=greenfill
    workbook.save(file)


def fillRedcolor(file,sheetname,rownum,colnum):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    redfill=PatternFill(start_color='ff0000',
                          end_color='ff0000',
                          fill_type='solid')
    sheet.cell(rownum,colnum).fill=redfill
    workbook.save(file)


import os
print(os.getcwd())